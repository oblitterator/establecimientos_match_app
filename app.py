"""
match_app — backend Flask para el vinculador REFES
Correr con: python app.py
Requiere: flask pandas openpyxl
"""
import subprocess, sys

def _ensure(*pkgs):
    for pkg in pkgs:
        try:
            __import__(pkg.split('[')[0].replace('-', '_'))
        except ImportError:
            print(f'Instalando {pkg}...')
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '--quiet'])

_ensure('flask', 'pandas', 'openpyxl', 'requests')

import os, re, uuid, unicodedata, threading
from pathlib import Path
import pandas as pd
from flask import Flask, request, jsonify, send_file, send_from_directory

import georef_normalizer as _geo

BASE_DIR = Path(__file__).resolve().parent

app = Flask(__name__, static_folder=str(BASE_DIR / 'static'), static_url_path='')

UPLOAD_DIR = BASE_DIR / 'uploads'
OUTPUT_DIR = BASE_DIR / 'outputs'
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

ALLOWED = {'.xlsx', '.xls', '.csv'}

BONUS_GEO_EXACTA = 15
BONUS_DOM_EXACTO = 10

# Números romanos ordenados de mayor a menor (mayor primero por si acaso,
# aunque los \b en regex hacen innecesario el orden estricto)
ROMAN_MAP = {
    'XVIII': 18, 'XVII': 17, 'XVI': 16, 'XV': 15, 'XIV': 14,
    'XIII': 13, 'XII': 12, 'XI': 11, 'X': 10, 'IX': 9,
    'VIII': 8, 'VII': 7, 'VI': 6, 'V': 5, 'IV': 4,
    'III': 3, 'II': 2, 'I': 1,
}

# ─────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────

def read_file(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext == '.csv':
        for enc in ['utf-8', 'latin-1', 'iso-8859-1']:
            try:
                return pd.read_csv(path, dtype=str, encoding=enc).fillna('')
            except UnicodeDecodeError:
                continue
    return pd.read_excel(path, dtype=str).fillna('')


def norm(s: str) -> str:
    if not s or str(s).strip() in ('nan', 'None', ''):
        return ''
    s = str(s).upper().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def strip_prefix(s: str, words: list) -> str:
    """Elimina prefijos del inicio. El separador puede ser espacio o punto."""
    for w in words:
        s = re.sub(r'^' + re.escape(w) + r'[\s\.]+', '', s, flags=re.IGNORECASE)
    return s.strip()


def replace_roman(s: str) -> str:
    """Reemplaza tokens de números romanos (palabras completas) por arábigos."""
    for roman, arabic in ROMAN_MAP.items():
        s = re.sub(r'\b' + roman + r'\b', str(arabic), s)
    return s


def to_indec(v, n: int):
    try:
        return str(int(float(v))).zfill(n)
    except (ValueError, TypeError):
        return None


def clean_cuit(s) -> str | None:
    if not s:
        return None
    c = re.sub(r'[^0-9]', '', str(s))
    return c if len(c) >= 10 else None


def validar_cuit(s) -> bool:
    """Valida dígito verificador de CUIT con módulo 11."""
    digits = re.sub(r'[^0-9]', '', str(s or ''))
    if len(digits) != 11:
        return False
    factors = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    total = sum(int(d) * f for d, f in zip(digits[:10], factors))
    resto = total % 11
    v = 0 if resto == 0 else (9 if resto == 1 else 11 - resto)
    return int(digits[10]) == v


def token_set_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    sa, sb = set(a.split()), set(b.split())
    inter = len(sa & sb)
    union = len(sa | sb)
    return round(inter / union * 100, 1) if union else 0.0


def exact_code(a, b) -> float:
    if not a or not b:
        return 0.0
    return 100.0 if str(a) == str(b) else 0.0


def dom_score(da: str, db: str) -> float:
    """Score de domicilio con bonus cuando comparten número de puerta."""
    base = token_set_ratio(da, db)
    nums_a = set(re.findall(r'\b\d+\b', da))
    nums_b = set(re.findall(r'\b\d+\b', db))
    if nums_a & nums_b and base < 50:
        base = min(100.0, base + 30)
    return base


# ─────────────────────────────────────────────────────────────────────
# SCORING
# ─────────────────────────────────────────────────────────────────────

def score_pair(ext_row: dict, ref_row: dict, cfg: dict) -> dict:
    cm_e = cfg['col_ext']
    cm_r = cfg['col_refes']
    name_pfx = cfg['name_prefixes']
    addr_pfx = cfg['addr_prefixes']
    W = cfg['weights']

    def gf(row, col):
        return row.get(col, '') if col else ''

    # Provincia
    ep = to_indec(gf(ext_row, cm_e.get('prov_indec')), 2)
    rp = to_indec(gf(ref_row, cm_r.get('prov_indec')), 2)
    if ep and rp:
        s_prov, pm = exact_code(ep, rp), 'codigo'
    else:
        s_prov, pm = token_set_ratio(
            norm(gf(ext_row, cm_e.get('prov_text'))),
            norm(gf(ref_row, cm_r.get('prov_text')))
        ), 'texto'

    # Filtro duro: provincias sin ninguna coincidencia → todo cero
    if s_prov == 0.0:
        return {
            'score_provincia':    0.0,
            'score_departamento': 0.0,
            'score_localidad':    0.0,
            'score_nombre':       0.0,
            'score_domicilio':    0.0,
            'score_total':        0.0,
            'prov_method': pm,
            'dept_method': 'n/a',
        }

    # Departamento
    ed = to_indec(gf(ext_row, cm_e.get('dept_indec')), 3)
    rd = to_indec(gf(ref_row, cm_r.get('dept_indec')), 3)
    if ed and rd:
        s_dept, dm = exact_code(ed, rd), 'codigo'
    else:
        s_dept, dm = token_set_ratio(
            norm(gf(ext_row, cm_e.get('dept_text'))),
            norm(gf(ref_row, cm_r.get('dept_text')))
        ), 'texto'

    # Localidad
    s_loc = token_set_ratio(
        norm(gf(ext_row, cm_e.get('loc'))),
        norm(gf(ref_row, cm_r.get('loc')))
    )

    # Nombre: quitar prefijo + normalizar romanos antes de comparar
    nom_e = replace_roman(strip_prefix(norm(gf(ext_row, cm_e.get('name'))), name_pfx))
    nom_r = replace_roman(strip_prefix(norm(gf(ref_row, cm_r.get('name'))), name_pfx))
    s_nom = token_set_ratio(nom_e, nom_r)

    # Domicilio
    if cfg.get('ext_addr_type') == 'split':
        dom_e = norm(gf(ext_row, cm_e.get('addr'))) + ' ' + norm(gf(ext_row, cm_e.get('num')))
    else:
        dom_e = norm(gf(ext_row, cm_e.get('addr')))

    if cfg.get('refes_addr_type') == 'split':
        dom_r = norm(gf(ref_row, cm_r.get('addr'))) + ' ' + norm(gf(ref_row, cm_r.get('num')))
    else:
        dom_r = norm(gf(ref_row, cm_r.get('addr')))

    da = strip_prefix(dom_e.strip(), addr_pfx)
    db = strip_prefix(dom_r.strip(), addr_pfx)
    s_dom = dom_score(da, db)

    # CABA (código INDEC 02): ignorar localidad y redistribuir su peso 40% NOM / 60% DOM
    prov_code_e = to_indec(gf(ext_row, cm_e.get('prov_indec', '')), 2)
    prov_name_e = norm(gf(ext_row, cm_e.get('prov_text', '')))
    is_caba = prov_code_e == '02' or any(
        kw in prov_name_e for kw in ['CAPITAL FEDERAL', 'CABA', 'CIUDAD AUTONOMA']
    )
    ew = dict(W)
    if is_caba:
        loc_w = ew['LOC']
        nom_bonus = round(loc_w * 0.4)
        ew['NOM'] = ew['NOM'] + nom_bonus
        ew['DOM'] = ew['DOM'] + (loc_w - nom_bonus)
        ew['LOC'] = 0
        s_loc = 0.0

    total = (
        s_prov * ew['PROV'] +
        s_dept * ew['DEPT'] +
        s_loc  * ew['LOC']  +
        s_nom  * ew['NOM']  +
        s_dom  * ew['DOM']
    ) / 100

    # Bonus: geo exacta por código
    if pm == 'codigo' and dm == 'codigo' and s_prov == 100.0 and s_dept == 100.0:
        total = min(100.0, total + BONUS_GEO_EXACTA)

    # Bonus: domicilio exacto
    if s_dom == 100.0:
        total = min(100.0, total + BONUS_DOM_EXACTO)

    return {
        'score_provincia':    round(s_prov, 1),
        'score_departamento': round(s_dept, 1),
        'score_localidad':    round(s_loc,  1),
        'score_nombre':       round(s_nom,  1),
        'score_domicilio':    round(s_dom,  1),
        'score_total':        round(total,  1),
        'prov_method': pm,
        'dept_method': dm,
    }


def greedy_match(ext_list: list, ref_list: list, cfg: dict,
                 id_col_ext: str, id_col_ref: str, thr_min: int = 30) -> dict:
    """
    Asignación exclusiva 1-a-1 estricta.
    - REFES con empate en su mejor score → bloqueadas (nadie las gana).
    - Pares con score < thr_min → descartados.
    - Sin segunda ronda: externos sin REFES exclusiva → SIN_MATCH.
    Retorna dict { ext_id: {'ref': ref_row, 'sc': score_dict} }
    """
    pairs = []
    for e in ext_list:
        for r in ref_list:
            sc = score_pair(e, r, cfg)
            pairs.append((e[id_col_ext], r[id_col_ref], sc, e, r))

    pairs.sort(key=lambda x: -x[2]['score_total'])

    # Detectar empates: para cada REFES, ¿cuántos externos obtienen su mejor score?
    best_for_ref: dict[str, float] = {}
    tie_count: dict[str, int] = {}
    for eid, rid, sc, e, r in pairs:
        s = sc['score_total']
        if rid not in best_for_ref or s > best_for_ref[rid]:
            best_for_ref[rid] = s
            tie_count[rid] = 1
        elif s == best_for_ref[rid]:
            tie_count[rid] += 1

    blocked_refs = {rid for rid, cnt in tie_count.items() if cnt > 1}

    used_e: set = set()
    used_r: set = set()
    matched: dict = {}

    for eid, rid, sc, e, r in pairs:
        if eid in used_e or rid in used_r:
            continue
        if rid in blocked_refs:
            continue
        if sc['score_total'] < thr_min:
            continue
        matched[eid] = {'ref': r, 'sc': sc}
        used_e.add(eid)
        used_r.add(rid)

    # Sin segunda ronda: externos no asignados quedan sin match
    return matched


def run_matching(ext_df: pd.DataFrame, ref_df: pd.DataFrame, cfg: dict) -> list:
    cm_e = cfg['col_ext']
    cm_r = cfg['col_refes']
    thr_good = cfg['thr_good']
    thr_warn = cfg['thr_warn']
    thr_min  = cfg.get('thr_min', 30)
    use_cuit = cfg.get('use_cuit', True)

    id_col_e = cm_e['id']
    id_col_r = cm_r['id']

    ext_rows = ext_df.to_dict('records')
    ref_rows = ref_df.to_dict('records')

    # ── Detección de anomalías en el archivo externo ──────────────────
    id_count: dict[str, int] = {}
    for e in ext_rows:
        eid = str(e.get(id_col_e, '') or '')
        if eid:
            id_count[eid] = id_count.get(eid, 0) + 1

    cuit_col_e = cm_e.get('cuit', '')
    for e in ext_rows:
        flags = []
        cuit_raw = e.get(cuit_col_e, '') if cuit_col_e else ''
        cuit_digits = re.sub(r'[^0-9]', '', str(cuit_raw or ''))
        if len(cuit_digits) < 10:
            flags.append('sin_cuit')
        elif len(cuit_digits) == 11 and not validar_cuit(cuit_raw):
            flags.append('cuit_invalido')

        eid = str(e.get(id_col_e, '') or '')
        if not eid:
            flags.append('id_nulo')
        elif id_count.get(eid, 0) > 1:
            flags.append('id_duplicado')

        e['__anomalia__'] = '|'.join(flags)

    # ── Indexar REFES por CUIT ────────────────────────────────────────
    ref_by_cuit: dict[str, list] = {}
    if use_cuit and cm_r.get('cuit'):
        for r in ref_rows:
            c = clean_cuit(r.get(cm_r['cuit'], ''))
            if c:
                ref_by_cuit.setdefault(c, []).append(r)

    # ── Indexar REFES por (prov_indec, dept_indec) ────────────────────
    ref_by_dept: dict[str, list] = {}
    for r in ref_rows:
        pv = to_indec(r.get(cm_r.get('prov_indec', ''), ''), 2)
        dv = to_indec(r.get(cm_r.get('dept_indec', ''), ''), 3)
        if pv and dv:
            k = f'{pv}-{dv}'
            ref_by_dept.setdefault(k, []).append(r)

    results = []

    def categorize(sc):
        if sc is None:
            return 'SIN_MATCH'
        if sc >= thr_good:
            return 'MATCH_BUENO'
        if sc >= thr_warn:
            return 'MATCH_REVISAR'
        return 'MATCH_DUDOSO'

    def build_row(e, ref, sc_dict, camino):
        sc_total = sc_dict['score_total'] if sc_dict else None
        return {
            'ext_id':              e.get(id_col_e, ''),
            'ext_nombre':          e.get(cm_e.get('name', ''), ''),
            'ext_provincia':       e.get(cm_e.get('prov_text', ''), '') or e.get(cm_e.get('prov_indec', ''), ''),
            'ext_departamento':    e.get(cm_e.get('dept_text', ''), '') or e.get(cm_e.get('dept_indec', ''), ''),
            'ext_localidad':       e.get(cm_e.get('loc', ''), ''),
            'ext_domicilio':       e.get(cm_e.get('addr', ''), ''),
            'ext_numero':          e.get(cm_e.get('num', ''), '') if cfg.get('ext_addr_type') == 'split' and cm_e.get('num') else None,
            'ext_lat':             e.get(cm_e.get('lat', ''), '') if cm_e.get('lat') else None,
            'ext_lon':             e.get(cm_e.get('lon', ''), '') if cm_e.get('lon') else None,
            'refes_id':            ref.get(id_col_r, '') if ref else None,
            'refes_nombre':        ref.get(cm_r.get('name', ''), '') if ref else None,
            'refes_provincia':     ref.get(cm_r.get('prov_text', ''), '') if ref else None,
            'refes_domicilio':     ref.get(cm_r.get('addr', ''), '') if ref else None,
            'refes_lat':           ref.get(cm_r.get('lat', ''), '') if ref and cm_r.get('lat') else None,
            'refes_lon':           ref.get(cm_r.get('lon', ''), '') if ref and cm_r.get('lon') else None,
            'score_provincia':     sc_dict['score_provincia']    if sc_dict else None,
            'score_departamento':  sc_dict['score_departamento'] if sc_dict else None,
            'score_localidad':     sc_dict['score_localidad']    if sc_dict else None,
            'score_nombre':        sc_dict['score_nombre']       if sc_dict else None,
            'score_domicilio':     sc_dict['score_domicilio']    if sc_dict else None,
            'score_total':         sc_total,
            'prov_method':         sc_dict['prov_method'] if sc_dict else None,
            'dept_method':         sc_dict['dept_method'] if sc_dict else None,
            'camino':              camino,
            'categoria':           categorize(sc_total),
            'tenia_cuit':          camino == 'A',
            'anomalia':            e.get('__anomalia__', ''),
        }

    # ── CAMINO A: agrupar por CUIT ────────────────────────────────────
    ext_a: dict[str, list] = {}
    ext_b: list = []

    for e in ext_rows:
        c = clean_cuit(e.get(cm_e.get('cuit', ''), '')) if use_cuit and cm_e.get('cuit') else None
        if c and c in ref_by_cuit:
            ext_a.setdefault(c, []).append(e)
        else:
            ext_b.append(e)

    for cuit, e_list in ext_a.items():
        r_list = ref_by_cuit[cuit]
        matched = greedy_match(e_list, r_list, cfg, id_col_e, id_col_r, thr_min)
        for e in e_list:
            eid = e[id_col_e]
            m = matched.get(eid)
            results.append(build_row(e, m['ref'] if m else None, m['sc'] if m else None, 'A'))

    # ── CAMINO B: agrupar por departamento INDEC ─────────────────────
    ext_b_by_dept: dict[str, list] = {}
    ext_b_no_indec: list = []

    for e in ext_b:
        pv = to_indec(e.get(cm_e.get('prov_indec', ''), ''), 2)
        dv = to_indec(e.get(cm_e.get('dept_indec', ''), ''), 3)
        if pv and dv:
            ext_b_by_dept.setdefault(f'{pv}-{dv}', []).append(e)
        else:
            ext_b_no_indec.append(e)

    for k, e_list in ext_b_by_dept.items():
        r_list = ref_by_dept.get(k, [])
        if not r_list:
            for e in e_list:
                results.append(build_row(e, None, None, 'B'))
        else:
            matched = greedy_match(e_list, r_list, cfg, id_col_e, id_col_r, thr_min)
            for e in e_list:
                eid = e[id_col_e]
                m = matched.get(eid)
                results.append(build_row(e, m['ref'] if m else None, m['sc'] if m else None, 'B'))

    for e in ext_b_no_indec:
        results.append(build_row(e, None, None, 'B'))

    return results


# ─────────────────────────────────────────────────────────────────────
# RUTAS API
# ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory(str(BASE_DIR / 'static'), 'index.html')


@app.route('/api/upload', methods=['POST'])
def upload():
    role = request.form.get('role')
    if 'file' not in request.files:
        return jsonify({'error': 'No se recibió archivo'}), 400

    f = request.files['file']
    ext = Path(f.filename).suffix.lower()
    if ext not in ALLOWED:
        return jsonify({'error': f'Formato no soportado: {ext}'}), 400

    fid = str(uuid.uuid4())
    dest = UPLOAD_DIR / f'{fid}{ext}'
    f.save(dest)

    try:
        df = read_file(dest)
    except Exception as e:
        return jsonify({'error': f'Error al leer el archivo: {e}'}), 400

    return jsonify({
        'file_id': fid,
        'filename': f.filename,
        'rows': len(df),
        'columns': list(df.columns),
    })


@app.route('/api/match', methods=['POST'])
def match():
    body = request.json

    try:
        ref_ext = next(p.suffix for p in UPLOAD_DIR.iterdir() if body['refes_id'] in p.name)
        ext_ext = next(p.suffix for p in UPLOAD_DIR.iterdir() if body['ext_id'] in p.name)
        ref_df  = read_file(UPLOAD_DIR / f"{body['refes_id']}{ref_ext}")
        ext_df  = read_file(UPLOAD_DIR / f"{body['ext_id']}{ext_ext}")
    except Exception as e:
        return jsonify({'error': f'Error cargando archivos: {e}'}), 400

    cfg = {
        'col_refes':       body['col_refes'],
        'col_ext':         body['col_ext'],
        'refes_addr_type': body.get('refes_addr_type', 'combined'),
        'ext_addr_type':   body.get('ext_addr_type', 'combined'),
        'weights': {
            'PROV': body['weights']['PROV'],
            'DEPT': body['weights']['DEPT'],
            'LOC':  body['weights']['LOC'],
            'NOM':  body['weights']['NOM'],
            'DOM':  body['weights']['DOM'],
        },
        'thr_good':      body.get('thr_good', 70),
        'thr_warn':      body.get('thr_warn', 50),
        'thr_min':       body.get('thr_min', 30),
        'use_cuit':      body.get('use_cuit', True),
        'name_prefixes': [norm(p) for p in body.get('name_prefixes', ['FARMACIA', 'FCIA', 'FARM'])],
        'addr_prefixes': [norm(p) for p in body.get('addr_prefixes', ['AVENIDA', 'AV', 'CALLE'])],
    }

    try:
        results = run_matching(ext_df, ref_df, cfg)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

    out_id = str(uuid.uuid4())
    _save_xlsx(results, OUTPUT_DIR / f'{out_id}.xlsx')

    def cnt(cat): return sum(1 for r in results if r['categoria'] == cat)

    scored = [r['score_total'] for r in results if r['score_total'] is not None]
    avg    = round(sum(scored) / len(scored), 1) if scored else None

    return jsonify({
        'out_id':    out_id,
        'total':     len(results),
        'bueno':     cnt('MATCH_BUENO'),
        'revisar':   cnt('MATCH_REVISAR'),
        'dudoso':    cnt('MATCH_DUDOSO'),
        'sin_match': cnt('SIN_MATCH'),
        'score_avg': avg,
        'results':   results,
    })


@app.route('/api/download/<out_id>')
def download(out_id):
    path = OUTPUT_DIR / f'{out_id}.xlsx'
    if not path.exists():
        return jsonify({'error': 'Archivo no encontrado'}), 404
    return send_file(path, as_attachment=True, download_name='match_refes_resultados.xlsx')


# ─────────────────────────────────────────────────────────────────────
# GEOREF — enriquecimiento de domicilios con API Georef
# ─────────────────────────────────────────────────────────────────────

GEO_JOBS: dict = {}
_GEO_LOCK = threading.Lock()


def _run_geo_job(job_id: str, df: pd.DataFrame, col_map: dict, addr_type: str, max_workers: int):
    """Corre en un thread aparte. Actualiza GEO_JOBS[job_id] con el progreso."""
    n_total = len(df)
    with _GEO_LOCK:
        GEO_JOBS[job_id]['total'] = n_total

    def progress_cb(done, total):
        with _GEO_LOCK:
            GEO_JOBS[job_id]['processed'] = done

    try:
        results = _geo.procesar_dataframe(
            df,
            col_provincia    = col_map.get('provincia')    or None,
            col_departamento = col_map.get('departamento') or None,
            col_domicilio    = col_map.get('domicilio')    or None,
            col_calle        = col_map.get('calle')        or None,
            col_numero       = col_map.get('numero')       or None,
            col_localidad    = col_map.get('localidad')    or None,
            addr_type        = addr_type,
            max_workers      = max_workers,
            progress_cb      = progress_cb,
        )

        out_path = OUTPUT_DIR / f'geo_{job_id}.xlsx'
        _save_geo_xlsx(results, out_path)

        n_ok  = sum(1 for r in results if r and r.get('latitud') is not None)
        n_err = sum(1 for r in results if r and r.get('domicilio_error'))

        with _GEO_LOCK:
            GEO_JOBS[job_id].update({
                'status':    'done',
                'processed': n_total,
                'con_punto': n_ok,
                'errores':   n_err,
                'out_path':  str(out_path),
            })

    except Exception as e:
        import traceback
        with _GEO_LOCK:
            GEO_JOBS[job_id].update({
                'status':    'error',
                'error_msg': str(e),
                'trace':     traceback.format_exc(),
            })


def _save_geo_xlsx(results: list, path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    if not results:
        pd.DataFrame().to_excel(path, index=False)
        return

    GEO_COLS = [
        'provincia_normalizada', 'id_provincia_indec',
        'departamento_normalizado', 'id_departamento_indec',
        'domicilio_normalizado', 'latitud', 'longitud',
        'provincia_error', 'departamento_error', 'domicilio_error',
    ]

    orig_cols = [c for c in results[0].keys() if c not in GEO_COLS]
    all_cols  = orig_cols + GEO_COLS

    wb = Workbook()
    ws = wb.active
    ws.title = 'Domicilios normalizados'

    fill_hdr  = PatternFill('solid', start_color='1F4E79')
    fill_ok   = PatternFill('solid', start_color='C6EFCE')
    fill_err  = PatternFill('solid', start_color='FFC7CE')
    fill_sep  = PatternFill('solid', start_color='DEEAF1')

    ws.append(all_cols)
    for i, c in enumerate(all_cols, 1):
        cell = ws.cell(1, i)
        cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        cell.fill      = fill_hdr if c not in GEO_COLS else fill_sep
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    lat_idx = all_cols.index('latitud') + 1 if 'latitud' in all_cols else None
    err_idx = all_cols.index('domicilio_error') + 1 if 'domicilio_error' in all_cols else None

    for r in results:
        if r is None:
            continue
        ws.append([r.get(c) for c in all_cols])
        row_n = ws.max_row
        if lat_idx:
            lat_val = ws.cell(row_n, lat_idx).value
            fill    = fill_ok if lat_val is not None else (fill_err if r.get('domicilio_error') else None)
            if fill:
                ws.cell(row_n, lat_idx).fill = fill
        if err_idx and r.get('domicilio_error'):
            ws.cell(row_n, err_idx).fill = fill_err

    for i, c in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = (
            22 if c in ('domicilio_normalizado', 'provincia_normalizada', 'departamento_normalizado')
            else 14 if c in ('latitud', 'longitud', 'id_departamento_indec', 'id_provincia_indec')
            else 28 if 'error' in c
            else 16
        )
    ws.freeze_panes = 'A2'

    wb.save(path)


@app.route('/geo/process', methods=['POST'])
def geo_process():
    body = request.json or {}
    file_id  = body.get('file_id')
    col_map  = body.get('col_map', {})
    addr_type    = body.get('addr_type', 'combined')
    max_workers  = min(int(body.get('max_workers', 5)), 10)

    if not file_id:
        return jsonify({'error': 'Falta file_id'}), 400

    # Buscar el archivo subido
    matches = list(UPLOAD_DIR.glob(f'{file_id}.*'))
    if not matches:
        return jsonify({'error': 'Archivo no encontrado'}), 404

    try:
        df = read_file(matches[0])
    except Exception as e:
        return jsonify({'error': f'Error al leer el archivo: {e}'}), 400

    job_id = str(uuid.uuid4())
    with _GEO_LOCK:
        GEO_JOBS[job_id] = {
            'status':    'running',
            'processed': 0,
            'total':     len(df),
            'con_punto': 0,
            'errores':   0,
            'out_path':  None,
            'error_msg': None,
        }

    t = threading.Thread(
        target=_run_geo_job,
        args=(job_id, df, col_map, addr_type, max_workers),
        daemon=True,
    )
    t.start()

    return jsonify({'job_id': job_id, 'total': len(df)})


@app.route('/geo/status/<job_id>')
def geo_status(job_id):
    with _GEO_LOCK:
        job = GEO_JOBS.get(job_id)
    if job is None:
        return jsonify({'error': 'Job no encontrado'}), 404
    return jsonify(job)


@app.route('/geo/download/<job_id>')
def geo_download(job_id):
    with _GEO_LOCK:
        job = GEO_JOBS.get(job_id)
    if not job or job['status'] != 'done':
        return jsonify({'error': 'Job no listo'}), 404
    path = Path(job['out_path'])
    if not path.exists():
        return jsonify({'error': 'Archivo no encontrado'}), 404
    return send_file(path, as_attachment=True, download_name='domicilios_normalizados.xlsx')


# ─────────────────────────────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────────────────────────────

def _write_data_sheet(ws, rows, cols, cat_fill, fill_hdr, fill_anom=None):
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    ws.append(cols)
    for i, c in enumerate(cols, 1):
        cell = ws.cell(1, i)
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    cat_idx   = (cols.index('categoria') + 1)   if 'categoria'   in cols else None
    score_idx = (cols.index('score_total') + 1) if 'score_total' in cols else None
    anom_idx  = (cols.index('anomalia') + 1)    if 'anomalia'    in cols else None

    orden = {'MATCH_BUENO': 0, 'MATCH_REVISAR': 1, 'MATCH_DUDOSO': 2, 'SIN_MATCH': 3}
    sorted_rows = sorted(rows, key=lambda r: (
        orden.get(r.get('categoria'), 4), -(r.get('score_total') or 0)
    ))

    for r in sorted_rows:
        ws.append([r.get(c) for c in cols])
        row_n = ws.max_row
        f = cat_fill.get(r.get('categoria'))
        if f:
            if cat_idx:
                ws.cell(row_n, cat_idx).fill = f
            if score_idx:
                ws.cell(row_n, score_idx).fill = f
        if anom_idx and fill_anom and r.get('anomalia'):
            ws.cell(row_n, anom_idx).fill = fill_anom

    widths = {
        'ext_id': 12, 'ext_nombre': 25, 'ext_provincia': 16, 'ext_departamento': 18,
        'ext_localidad': 18, 'ext_domicilio': 22, 'ext_numero': 8, 'ext_lat': 12, 'ext_lon': 12,
        'refes_id': 18, 'refes_nombre': 25, 'refes_provincia': 16, 'refes_domicilio': 22,
        'refes_lat': 12, 'refes_lon': 12,
        'score_provincia': 13, 'score_departamento': 14, 'score_localidad': 13,
        'score_nombre': 12, 'score_domicilio': 13, 'score_total': 11,
        'prov_method': 10, 'dept_method': 10, 'camino': 8, 'categoria': 18,
        'tenia_cuit': 10, 'anomalia': 24,
    }
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 13)
    ws.freeze_panes = 'A2'


def _save_xlsx(results: list, path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    fill_hdr    = PatternFill('solid', start_color='1F4E79')
    fill_green  = PatternFill('solid', start_color='C6EFCE')
    fill_yellow = PatternFill('solid', start_color='FFEB9C')
    fill_red    = PatternFill('solid', start_color='FFC7CE')
    fill_gray   = PatternFill('solid', start_color='D9D9D9')
    fill_anom   = PatternFill('solid', start_color='FFF3CD')

    cat_fill = {
        'MATCH_BUENO':   fill_green,
        'MATCH_REVISAR': fill_yellow,
        'MATCH_DUDOSO':  fill_red,
        'SIN_MATCH':     fill_gray,
    }

    base_cols = [
        'ext_id', 'ext_nombre', 'ext_provincia', 'ext_departamento', 'ext_localidad',
        'ext_domicilio', 'ext_numero', 'ext_lat', 'ext_lon',
        'refes_id', 'refes_nombre', 'refes_provincia', 'refes_domicilio',
        'refes_lat', 'refes_lon',
        'score_provincia', 'score_departamento', 'score_localidad',
        'score_nombre', 'score_domicilio', 'score_total',
        'prov_method', 'dept_method', 'camino', 'categoria',
    ]
    cols_sin = base_cols + ['tenia_cuit', 'anomalia']

    wb = Workbook()

    # Hoja 1: Con match (MATCH_BUENO + MATCH_REVISAR)
    ws1 = wb.active
    ws1.title = 'Con match'
    con_match = [r for r in results if r.get('categoria') in ('MATCH_BUENO', 'MATCH_REVISAR')]
    _write_data_sheet(ws1, con_match, base_cols, cat_fill, fill_hdr)

    # Hoja 2: Sin match (resto)
    ws2 = wb.create_sheet('Sin match')
    sin_match = [r for r in results if r.get('categoria') not in ('MATCH_BUENO', 'MATCH_REVISAR')]
    _write_data_sheet(ws2, sin_match, cols_sin, cat_fill, fill_hdr, fill_anom=fill_anom)

    # Hoja 3: Resumen (los totales deben sumar el total del externo)
    ws3 = wb.create_sheet('Resumen')
    ws3.append(['Categoría', 'Cantidad'])
    for cat, label in [
        ('MATCH_BUENO',   'Match bueno'),
        ('MATCH_REVISAR', 'Revisar'),
        ('MATCH_DUDOSO',  'Dudoso'),
        ('SIN_MATCH',     'Sin match'),
    ]:
        n = sum(1 for r in results if r.get('categoria') == cat)
        ws3.append([label, n])
    ws3.append(['TOTAL', len(results)])

    # Estilo mínimo para la hoja resumen
    from openpyxl.styles import Font as OFont
    for cell in ws3[1]:
        cell.font = OFont(bold=True)

    wb.save(path)


if __name__ == '__main__':
    print('Iniciando servidor en http://localhost:5001')
    app.run(debug=True, port=5001)
